import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from db import init_db, get_all_flight_ids, get_combinations_for_flight, get_prices_for_combination

CONFIG_PATH = "config.json"
OUTPUT_PATH = "flight_prices.xlsx"

# Colours
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"
BLACK      = "000000"
BORDER_CLR = "BFBFBF"
GREEN_BG   = "E2EFDA"

def thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=10, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def bfont(size=10, bold=False, color=BLACK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)

def load_config():
    with open(CONFIG_PATH) as f:
        return json.load(f)


def write_cell(ws, row, col, value, font=None, fill_color=None, num_fmt=None, align="center", bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or bfont(bold=bold)
    if fill_color:
        c.fill = fill(fill_color)
    c.border = thin_border()
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_fmt:
        c.number_format = num_fmt
    return c


def build_flight_sheet(wb, flight_cfg, all_combinations):
    """
    One sheet per flight config.
    Columns: Date | Combination | Out Time | Out Airline | Ret Airline | Price | vs Previous
    Grouped by combination_id so each combo's history is visible side by side.
    """
    ws = wb.create_sheet(title=flight_cfg["id"][:31])

    # Title
    total_cols = 7
    ws.merge_cells(f"A1:G1")
    c = ws["A1"]
    c.value = flight_cfg["label"]
    c.font = hfont(size=13)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (f"Route: {flight_cfg['origin']} to {flight_cfg['destination']}   |   "
               f"Out: {flight_cfg['outbound_date']}   Return: {flight_cfg['return_date']}   |   "
               f"Outbound filter: after {flight_cfg.get('outbound_after','any')}")
    c.font = hfont(size=9)
    c.fill = fill(MID_BLUE)
    c.alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Date fetched", "Combination", "Out time", "Out airline", "Ret airline", "Price (EUR)", "vs Prev (EUR)"]
    col_widths = [14, 22, 10, 16, 16, 14, 16]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hfont(size=10)
        c.fill = fill(MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 5
    chart_data = {}  # combination_id -> list of (date_str, price)

    # Write data grouped by combination (filter out any None ids from old runs)
    for combo_id in sorted(c for c in all_combinations if c is not None):
        rows = get_prices_for_combination(flight_cfg["id"], combo_id)
        if not rows:
            continue

        # Group header row
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value=f"  Combination: {combo_id}")
        c.font = hfont(size=10, color=WHITE)
        c.fill = fill(MID_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        chart_data[combo_id] = []
        prev_price = None

        for i, (fetched_at, price, currency, out_time, out_airline,
                ret_time, ret_airline, stops, duration) in enumerate(rows):
            bg = GREY if i % 2 == 0 else WHITE
            dt = datetime.strptime(fetched_at, "%Y-%m-%d %H:%M:%S")
            date_str = dt.strftime("%Y-%m-%d")

            write_cell(ws, row, 1, date_str,         fill_color=bg)
            write_cell(ws, row, 2, combo_id,          fill_color=bg, align="left")
            write_cell(ws, row, 3, out_time or "?",   fill_color=bg)
            write_cell(ws, row, 4, out_airline or "?",fill_color=bg)
            write_cell(ws, row, 5, ret_airline or "?",fill_color=bg)
            write_cell(ws, row, 6, price,              fill_color=bg, num_fmt="EUR#,##0.00")

            # vs previous
            if prev_price is not None and price is not None:
                diff = price - prev_price
                vc = ws.cell(row=row, column=7, value=diff)
                vc.font = bfont(color="375623" if diff <= 0 else "C00000")
                vc.fill = fill(bg)
                vc.border = thin_border()
                vc.alignment = Alignment(horizontal="center")
                vc.number_format = 'EUR#,##0.00;[Red]-EUR#,##0.00'
            else:
                write_cell(ws, row, 7, "--", fill_color=bg)

            prev_price = price
            if price:
                chart_data[combo_id].append((date_str, price))
            row += 1

        row += 1  # blank row between combinations

    # Summary stats per combination
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="Summary — Cheapest price seen per combination")
    c.font = hfont(size=11)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    sum_headers = ["Combination", "Out time", "Out airline", "Ret airline", "Min price", "Max price", "Avg price"]
    for col, h in enumerate(sum_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont(size=10)
        c.fill = fill(MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
    row += 1

    for i, combo_id in enumerate(sorted(c for c in all_combinations if c is not None)):
        rows_data = get_prices_for_combination(flight_cfg["id"], combo_id)
        if not rows_data:
            continue
        prices = [r[1] for r in rows_data if r[1] is not None]
        out_time    = rows_data[-1][3] or "?"
        out_airline = rows_data[-1][4] or "?"
        ret_airline = rows_data[-1][6] or "?"
        bg = GREY if i % 2 == 0 else WHITE

        write_cell(ws, row, 1, combo_id,              fill_color=bg, align="left")
        write_cell(ws, row, 2, out_time,               fill_color=bg)
        write_cell(ws, row, 3, out_airline,             fill_color=bg)
        write_cell(ws, row, 4, ret_airline,             fill_color=bg)
        write_cell(ws, row, 5, min(prices) if prices else "N/A", fill_color=bg, num_fmt="EUR#,##0.00")
        write_cell(ws, row, 6, max(prices) if prices else "N/A", fill_color=bg, num_fmt="EUR#,##0.00")
        write_cell(ws, row, 7, round(sum(prices)/len(prices),2) if prices else "N/A", fill_color=bg, num_fmt="EUR#,##0.00")
        row += 1

    return ws


def build_summary_sheet(wb, config, all_ids):
    ws = wb.create_sheet(title="Summary", index=0)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Flight Price Tracker - Summary"
    c.font = hfont(size=14)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = hfont(size=9)
    c.fill = fill(MID_BLUE)
    c.alignment = Alignment(horizontal="center")

    headers = ["Flight", "Route", "Outbound date", "Return date", "# Combinations tracked", "Cheapest seen"]
    widths =  [28,       14,      16,               14,            24,                        16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hfont(size=10)
        c.fill = fill(MID_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    flight_map = {f["id"]: f for f in config["flights"]}

    for i, fid in enumerate(all_ids):
        row = 5 + i
        fcfg = flight_map.get(fid, {})
        bg = GREY if i % 2 == 0 else WHITE

        combos = get_combinations_for_flight(fid)
        all_prices = []
        for combo_id in combos:
            rows_data = get_prices_for_combination(fid, combo_id)
            all_prices += [r[1] for r in rows_data if r[1] is not None]

        cheapest = min(all_prices) if all_prices else "N/A"

        write_cell(ws, row, 1, fcfg.get("label", fid), fill_color=bg, align="left", bold=True)
        write_cell(ws, row, 2, f"{fcfg.get('origin','?')} to {fcfg.get('destination','?')}", fill_color=bg)
        write_cell(ws, row, 3, fcfg.get("outbound_date",""), fill_color=bg)
        write_cell(ws, row, 4, fcfg.get("return_date",""),   fill_color=bg)
        write_cell(ws, row, 5, len(combos),                   fill_color=bg)
        write_cell(ws, row, 6, cheapest, fill_color=bg,
                   num_fmt="EUR#,##0.00" if isinstance(cheapest, (int, float)) else None)


def run():
    init_db()
    config = load_config()
    flight_map = {f["id"]: f for f in config["flights"]}
    all_ids = [f["id"] for f in config["flights"]]

    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, config, all_ids)

    for fid in all_ids:
        fcfg = flight_map.get(fid)
        if not fcfg:
            continue
        combos = get_combinations_for_flight(fid)
        build_flight_sheet(wb, fcfg, combos)

    wb.save(OUTPUT_PATH)
    print(f"Dashboard saved -> {OUTPUT_PATH}")


if __name__ == "__main__":
    run()
